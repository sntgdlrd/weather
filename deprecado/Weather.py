import os
import openpyxl
from openpyxl import Workbook
import requests

from datetime import datetime, timedelta

def crear_o_abrir_excel(ruta_completa):
    try:
        # Intenta abrir el archivo existente
        wb = openpyxl.load_workbook(ruta_completa)
        print("El archivo Excel ya existe. Abriendo...")
    except FileNotFoundError:
        # Si el archivo no existe, crea uno nuevo
        wb = Workbook()
        wb.save(ruta_completa)
        print("El archivo Excel no existía. Se ha creado uno nuevo.")
    return wb

def busqueda_datos(listado, fecha):
    resultados_totales = []
    hoy = datetime.today()

    for item in listado:
        url = f'http://api.weatherapi.com/v1/history.json?key=d8cb833dde7a48daa8b131011242802&q={item["name"]}&q={item["region"]}&q={item["country"]}&dt={fecha}'
        try:
            response = requests.get(url)
            # Esto generará una excepción si hay un error HTTP
            response.raise_for_status()
        except requests.RequestException as e:
            print(f"No se pudo obtener la respuesta para {url}: {e}")
            continue
        
        moonrise = 'No' in response.json()['forecast']['forecastday'][0]['astro']['moonrise']
        
        moonset = 'No' in response.json()['forecast']['forecastday'][0]['astro']['moonset']

        # response.json()['location']['name'], response.json()['location']['region'], response.json()['location']['country'], 
        resultado = [item['id'], response.json()['forecast']['forecastday'][0]['date'], response.json()['forecast']['forecastday'][0]['day']['maxtemp_c'], response.json()['forecast']['forecastday'][0]['day']['mintemp_c'],
        response.json()['forecast']['forecastday'][0]['day']['maxwind_kph'], response.json()['forecast']['forecastday'][0]['day']['totalprecip_mm'], response.json()['forecast']['forecastday'][0]['day']['totalsnow_cm'],
        response.json()['forecast']['forecastday'][0]['day']['avgvis_km'], response.json()['forecast']['forecastday'][0]['day']['avghumidity'], response.json()['forecast']['forecastday'][0]['day']['daily_will_it_rain'],
        response.json()['forecast']['forecastday'][0]['day']['daily_chance_of_rain'], response.json()['forecast']['forecastday'][0]['day']['daily_will_it_snow'], response.json()['forecast']['forecastday'][0]['day']['daily_chance_of_snow'], 
        response.json()['forecast']['forecastday'][0]['day']['condition']['text'], response.json()['forecast']['forecastday'][0]['day']['uv'], response.json()['forecast']['forecastday'][0]['astro']['sunrise'], response.json()['forecast']['forecastday'][0]['astro']['sunset'],
        response.json()['forecast']['forecastday'][0]['astro']['moonrise'], moonrise, response.json()['forecast']['forecastday'][0]['astro']['moonset'], moonset, response.json()['forecast']['forecastday'][0]['astro']['moon_phase'], response.json()['forecast']['forecastday'][0]['astro']['moon_illumination'], hoy]

        resultados_totales.append(resultado)
    return(resultados_totales)




def agregar_datos(wb, sheet_name, column_names, regiones):
    
    # Obtener la fecha de hoy
    hoy = datetime.today()
    cantidad_dias = 365

    try:
        sheet = wb[sheet_name]
    except KeyError:
        # Si la hoja no existe, crear una nueva
        sheet = wb.create_sheet(title=sheet_name)
        sheet.append(column_names)  # Agrega los nombres de las columnas
    
    # Iterar desde hoy hasta 365 días atrás
    for i in range(cantidad_dias):
        # Calcular la fecha restando i días a la fecha actual
        fecha = hoy - timedelta(days=i)
        # Formatear la fecha en el formato YYYY-mm-dd
        fecha_formateada = fecha.strftime('%Y-%m-%d')
        # Imprimir la fecha formateada
        print(fecha_formateada)
        listado = busqueda_datos(regiones, fecha_formateada)
        #print(listado)

        for item in listado:
            sheet.append(item)


# Nombre del archivo Excel
nombre_archivo = "weather.xlsx"

# Obtiene el directorio del script en ejecución
directorio_script = os.path.dirname(os.path.abspath(__file__))
ruta_completa = os.path.join(directorio_script, nombre_archivo)

# Nombres de las hojas
nombre_hoja_arg = "forecastday_arg"
nombre_hoja_world = "forecastday_world"

# Nombres de las columnas
column_names = ['id', 'date', 'maxtemp_c', 'avgtemp_c', 'maxwind_kph', 'totalprecip_mm', 'totalsnow_cm', 'avgvis_km', 'avghumidity', 'daily_will_it_rain', 'daily_chance_of_rain', 'daily_will_it_snow', 'daily_chance_of_snow', 'condition', 'uv', 'sunrise', 'sunset', 'moonrise', 'had_moonrise', 'moonset', 'had_moonset', 'moon_phase', 'moon_illumination', 'entry_date']

# Crear o abrir el archivo Excel
wb = crear_o_abrir_excel(ruta_completa)

# Datos que quieres agregar
provincias = [{'id': 1 , 'name':'La Plata', 'region': 'Buenos Aires', 'country': 'Argentina'},
            {'id': 2 , 'name':'Districto Federal', 'region': 'Buenos Aires', 'country': 'Argentina'},
            {'id': 3 , 'name':'San Fernando del Valle de Catamarca', 'region': 'Catamarca', 'country': 'Argentina'},
            {'id': 4 , 'name':'Resistencia', 'region': 'Chaco', 'country': 'Argentina'},
            {'id': 5 , 'name':'Rawson','region': 'Chubut','country': 'Argentina'},
            {'id': 6 , 'name':'Córdoba','region': 'Córdoba','country': 'Argentina'},
            {'id': 7 , 'name':'Corrientes','region': 'Corrientes','country': 'Argentina'},
            {'id': 8 , 'name':'Paraná','region': 'Entre Ríos','country': 'Argentina'},
            {'id': 9 , 'name':'Formosa','region': 'Formosa','country': 'Argentina'},
            {'id': 10 , 'name':'San Salvador de Jujuy','region': 'Jujuy','country': 'Argentina'},
            {'id': 11 , 'name':'Santa Rosa','region': 'La Pampa','country': 'Argentina'},
            {'id': 12 , 'name':'La Rioja','region': 'La Rioja','country': 'Argentina'},
            {'id': 13 , 'name':'Mendoza','region': 'Mendoza','country': 'Argentina'},
            {'id': 14 , 'name':'Posadas','region': 'Misiones','country': 'Argentina'},
            {'id': 15 , 'name':'Neuquén','region': 'Neuquén','country': 'Argentina'},
            {'id': 16 , 'name':'Viedma','region': 'Río Negro','country': 'Argentina'},
            {'id': 17 , 'name':'Salta','region': 'Salta','country': 'Argentina'},
            {'id': 18 , 'name':'San Juan','region': 'San Juan','country': 'Argentina'},
            {'id': 19 , 'name':'San Luis','region': 'San Luis','country': 'Argentina'},
            {'id': 20 , 'name':'Río Gallegos','region': 'Santa Cruz','country': 'Argentina'},
            {'id': 21 , 'name':'Santa Fe','region': 'Santa Fe','country': 'Argentina'},
            {'id': 22 , 'name':'Santiago del Estero','region': 'Santiago del Estero','country': 'Argentina'},
            {'id': 23 , 'name':'Ushuaia','region': 'Tierra del Fuego','country': 'Argentina'},
            {'id': 24 , 'name':'San Miguel de Tucumán','region': 'Tucumán','country': 'Argentina'}]

ciudades = [{'id': 25 , 'name':'Washington', 'region': '', 'country': 'United States of America'},
            {'id': 26 , 'name':'Beijing', 'region': '', 'country': 'China'},
            {'id': 27 , 'name':'Tokyo', 'region': '', 'country': 'Japan'},
            {'id': 28 , 'name':'Berlin', 'region': '', 'country': 'Germany'},
            {'id': 29 , 'name':'New Delhi', 'region': '', 'country': 'India'},
            {'id': 30 , 'name':'London', 'region': '', 'country': 'United Kingdom'},
            {'id': 31 , 'name':'Paris', 'region': '', 'country': 'France'},
            {'id': 32 , 'name':'Rome', 'region': '', 'country': 'Italy'},
            {'id': 33 , 'name':'Brasilia', 'region': 'Distrito Federal', 'country': 'Brasil'},
            {'id': 34 , 'name':'Ottawa', 'region': '', 'country': 'Canada'},
            {'id': 35 , 'name':'Seoul', 'region': '', 'country': 'South Korea'},
            {'id': 36 , 'name':'Moscow', 'region': '', 'country': 'Russia'},
            {'id': 37 , 'name':'Canberra', 'region': '', 'country': 'Australia'},
            {'id': 38 , 'name':'Madrid', 'region': '', 'country': 'Spain'},
            {'id': 39 , 'name':'Mexico City', 'region': '', 'country': 'Mexico'}]

# Agregar datos a la hoja
agregar_datos(wb, nombre_hoja_arg, column_names, provincias)
agregar_datos(wb, nombre_hoja_world, column_names, ciudades)

# Guardar los cambios en el archivo Excel
wb.save(ruta_completa)
print("Datos agregados exitosamente.")